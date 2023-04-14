from leetcode_handler import LeetcodeHandler
from chatgpt_api import promt_template
from bs4 import BeautifulSoup
import openpyxl
import time

#refernce https://github.com/boolalpha/GPTLeetCode/blob/master/src/backend/CodeAccuracy/main.py
def get_leetcode_problems(leetcode_handler):
    # 1. Get all the problems from leetcode
    # get all the urls of the problem pages
    print("Getting all problem urls")
    problem_urls = leetcode_handler.get_all_problems_urls()
    print(problem_urls)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    headers = ["Problem number","Problem tittle", "Problem Content", "Acceptance","Difficulty", "Answers by ChatGPT","Succeeded","Runtime","runtime_beats", "memory","memory_beats","error_type","error_message","Total_testcases","Test_cases_passed"]
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    excel_data = []
    count = 0
    # 1.a. Iterate all the pages and insert them into the dB
    for url in problem_urls:
        # get all the problems posted on the problems page
        print("Scraping problems from page:\n\t%s", url)
        page_results = leetcode_handler.parse_problems(url)
        print(page_results, "page_results")
        for problem in page_results:
            time.sleep(1)
            (problem_header, problem_content) = leetcode_handler.parse_problem_content(problem['problem_url'], 'Java')
            
            # if there are topic tags we will add them to the table
            print("Problem URL {}".format(problem["problem_url"]))
                # get the inserted primary key
                # topic_tags_list = [[primary_key, topic] for topic in problem["topic_tags"]]
            try:
                openai_response = promt_template(problem_content=problem_content, header=problem_header)

            except:
                time.sleep(120)  # wait for 2 minutes before trying again
                openai_response = promt_template(problem_content=problem_content, header=problem_header)

            print(openai_response,"openai_response")
            result = leetcode_handler.submit_problem("Java", problem["problem_url"], openai_response)

            print(result)
            print("XXXXXXXxXXXXXXXXXXXXXXXXXXXXXXXX")
            if result["error_type"] is None:
                error_type = "None"
            else:
                error_type = result["error_type"]

            if result["error_message"] is None:
                errot_message = "None"
            else:
                errot_message = result["error_message"]

            if result["total_testcases"] is None:
                total_testcases = "None"
            else:
                total_testcases = result["total_testcases"]

            if result["testcases_passed"] is None:
                testcases_passed = "None"
            else:
                testcases_passed= result["testcases_passed"]

            question_data = [problem["problem_number"],problem["title"],problem_content,problem["acceptance"],problem["difficulty"], 
                             openai_response, result["succeeded"],result["runtime"],result["runtime_beats"],
                             result["memory"],result["memory_beats"],error_type, errot_message,total_testcases,testcases_passed]
            
            excel_data.append(question_data)

            if result["succeeded"]:
                print("111")
                # 创建一个文件，并将字符串写入文件中
                with open('Java/'+str(problem["problem_number"])+ '.class', 'w') as file:
                    file.write(openai_response)

            count += 1
            if count ==100:
                break
        if count == 100:
            break
        print("Done scraping problems from page:\n\t%s", url)

    for row, row_data in enumerate(excel_data, 2):
        for col, cell_data in enumerate(row_data, 1):
            worksheet.cell(row=row, column=col, value=cell_data)

        # Save the workbook
        workbook.save('data_J.xlsx')
        



def main():
    leetcode = LeetcodeHandler()
    leetcode.login()
    get_leetcode_problems(leetcode)


if __name__ == "__main__":

    # call to main
    main()